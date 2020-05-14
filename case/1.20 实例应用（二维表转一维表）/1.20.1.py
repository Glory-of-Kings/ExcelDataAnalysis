import openpyxl
wb=openpyxl.load_workbook('业绩表.xlsx')
ws=wb.worksheets[0]
if '转换表' in wb.sheetnames:
    wb.remove(wb['转换表'])
nws=wb.create_sheet('转换表');nws.append(['姓名','月份','业绩'])
rng1=ws['a'][1:]
rng2=ws.iter_rows(min_col=2,min_row=2)
for name,row in list(zip(rng1,rng2)):
    for x,y in zip(ws[1][1:],row):
        nws.append([name.value,x.value,y.value])
wb.save('业绩表.xlsx')