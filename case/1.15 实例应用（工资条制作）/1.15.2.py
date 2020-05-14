import openpyxl
wb=openpyxl.load_workbook('工资表.xlsx',data_only=True)
ws=wb.active
for r in range(ws.max_row,2,-1):
    ws.insert_rows(r,2)
    for c in range(1,8):
        ws.cell(r+1,c).value=ws.cell(1,c).value
wb.save('工资条2.xlsx')