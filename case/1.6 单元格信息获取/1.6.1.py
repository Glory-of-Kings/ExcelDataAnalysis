import openpyxl
wb=openpyxl.load_workbook('demo.xlsx')
ws=wb.worksheets[0]
print(ws['b1'].value)
print(ws.cell(1,2).value)
print(openpyxl.load_workbook('demo.xlsx').worksheets[0]['b1'].value)
