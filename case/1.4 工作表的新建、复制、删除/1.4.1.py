import openpyxl
# wb=openpyxl.Workbook()
# wb.create_sheet()
# wb.create_sheet()
# wb.create_sheet()
# wb.save('demo1.xlsx')

wb=openpyxl.load_workbook('demo2.xlsx')
wb.create_sheet('工资表',2)
wb.save('demo2.xlsx')