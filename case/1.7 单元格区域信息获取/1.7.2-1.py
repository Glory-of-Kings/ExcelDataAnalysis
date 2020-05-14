import openpyxl
import os
path = os.path.dirname(__file__)
wb=openpyxl.load_workbook( path + '/test.xlsx')
ws=wb.active
rngs=ws['a2:e71']
# print(['%s-%d'%(row[0].value,sum([c.value for c in row][1:])) for row in rngs])

# print([ '%s-%d'%(row[0].value ,sum(([c.value for c in row])[1:])) for row in rngs])
print( [[c.value for c in row] for row in ws['b:d']])