import openpyxl
import os
import sys
path = os.path.dirname(__file__)
wb=openpyxl.load_workbook( path + '/test.xlsx')
ws=wb.active
# print([[c.value for c in row] for row in ws['a1:d3']])
# print(list(ws.values)[1:4])
# print(ws['A1:B12'])
for row in ws['A1:B12']:
    for c in row:
        print(c.value)

print([[c.value for c in row ] for row in ws['A1:B12']])
print()
print(list(ws.values)[0:4])  



