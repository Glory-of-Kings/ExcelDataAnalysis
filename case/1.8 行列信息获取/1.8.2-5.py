import openpyxl
wb=openpyxl.load_workbook('demo.xlsx')
ws=wb.active
minr=ws.min_row
minc=ws.min_column
maxr=ws.max_row
maxc=ws.max_column
rngs=ws.iter_rows(min_row=minr+1,min_col=minc+2,max_row=maxr-1,max_col=maxc-1)
col=ws.iter_cols(min_row=minr+1,min_col=minc+1,max_row=maxr-1,max_col=minc+1)
total=([sum([v.value for v in row]) for row in rngs])
cp=([[v.value for v in row] for row in col][0])
print(list(zip(cp,total)))
