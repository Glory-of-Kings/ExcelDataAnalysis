import openpyxl
wb=openpyxl.load_workbook('test.xlsx')
ws=wb.worksheets[0]
total=[sum([c.value for c in col]) for col in ws.iter_cols(min_col=2,min_row=2)]
name=[c.value for c in ws[1][1:]]
print(list(zip(name,total)))