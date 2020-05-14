import openpyxl
wb=openpyxl.load_workbook('test.xlsx')
ws=wb.worksheets[0]
total=[sum([c.value for c in row]) for row in ws.iter_rows(min_col=2,min_row=2)]
name=[c.value for c in ws['a']][1:]
print(list(zip(name,total)))