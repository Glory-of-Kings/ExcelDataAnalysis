import openpyxl
wb=openpyxl.load_workbook('test.xlsx')
ws=wb.worksheets[0]
rngs=ws[2:ws.max_row]
nws=wb.create_sheet('结果')
nws.append(['姓名','总分'])
for row in rngs:
    nws.append([row[0].value,sum([c.value for c in row][1:])])
wb.save('test1.xlsx')