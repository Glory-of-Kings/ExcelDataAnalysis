import openpyxl
wb=openpyxl.load_workbook('成绩表.xlsx')
ws=wb.active
for r in range(ws.max_row,1,-1):
    s=sum([c.value for c in ws[r]][1:])
    if s>=300:
        ws.delete_rows(r)
wb.save('成绩表筛选结果.xlsx')