import openpyxl
wb=openpyxl.load_workbook('test.xlsx')
ws=wb.worksheets[1]
ws['a1']=123
ws.cell(2,3,'我是中国人')
ws.cell(3,3).value='我是四川人'
wb.save('test.xlsx')