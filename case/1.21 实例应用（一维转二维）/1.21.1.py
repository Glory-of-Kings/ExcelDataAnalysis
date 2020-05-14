import openpyxl
wb=openpyxl.load_workbook('业绩表.xlsx')
if not '二维表' in wb.sheetnames:
    nws=wb.create_sheet('二维表')
    ws=wb.worksheets[0]
    rngs=list(ws.values)[1:]
    mm=list({m.value:'' for m in ws['b'][1:]})
    name=list({m.value:'' for m in ws['a'][1:]})
    nws.append(['姓名']+mm)
    for n in name:
        l=[(n,m) for m in mm]
        nws.append([n]+[list(filter(lambda r:t[0]==r[0] and t[1]==r[1],rngs))[0][2] for t in l])
    wb.save('业绩表.xlsx')
