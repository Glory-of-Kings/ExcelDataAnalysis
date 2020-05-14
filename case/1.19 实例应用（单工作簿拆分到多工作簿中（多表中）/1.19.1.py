import openpyxl
wb=openpyxl.load_workbook('工资表.xlsx',data_only=True)
ws=wb.active
rngs=list(ws.values)
d={}
for row in rngs[1:]:
    if row[2] in d.keys() and row[3] in d[row[2]].keys():
        d[row[2]][row[3]]+= [row]
    else:
        if not row[2] in d.keys():d[row[2]]={}
        d[row[2]][row[3]]=[row]
for k,v in d.items():
    nwb=openpyxl.Workbook()
    for k1,v1 in v.items():
        nws=nwb.create_sheet(k1)
        nws.append(rngs[0])
        for v2 in v1:
            nws.append(v2)
    nwb.remove(nwb.active)
    nwb.save('拆分结果/'+k+'.xlsx')