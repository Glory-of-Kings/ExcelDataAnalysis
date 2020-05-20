'''
对问题评级，做数据分析，输出饼图 ------ 纪志昌

'''

import os
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.marker import DataPoint
from loggerhelper import *

PATH = os.path.dirname(__file__)  # 文件运行目录


def get_requestleval_piechart():
    print("当前文件路径:" + PATH)
    workbook = openpyxl.load_workbook(PATH + '/total.xlsx')
    worksheet_total = workbook['total']
    cols = worksheet_total['I']  # 获取I列问题评级
    levels = [c.value for c in cols]
    logger.info(levels)
    dic_level = {}
    for level in levels:
        level_name = level.split()
        if dic_level.get(level_name[0], 'None') != 'None':
            dic_level[level_name[0]] += 1
        else:
            dic_level.setdefault(level_name[0], 1)

    worksheet_level = workbook.create_sheet()  # 创建临时表
    worksheet_level.title = 'Request Level'
    dic_level_sorted = sorted(dic_level.items(), key=lambda x: x[1], reverse=True)
    # print(dic_level_sorted)
    worksheet_level.append(['级别', '个数'])
    for k in dic_level_sorted:
        worksheet_level.append([k[0], k[1]])
        logger.info([k[0], k[1]])
    add_pie_chart(workbook, worksheet_level)


def add_pie_chart(workbook, worksheet_level):
    pie = PieChart()
    cats = Reference(worksheet_level, min_col=1, min_row=2, max_row=worksheet_level.max_row)
    data = Reference(worksheet_level, min_col=2, min_row=1, max_row=worksheet_level.max_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)
    pie.title = '问题评级'
    data_points = DataPoint(idx=0, explosion=20)
    pie.series[0].data_points = [data_points]
    worksheet_level.add_chart(pie, "C9")
    workbook.save(PATH + '/total.xlsx')
