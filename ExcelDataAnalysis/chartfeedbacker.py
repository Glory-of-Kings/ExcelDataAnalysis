'''
对问题的反馈人，做数据分析，输出柱形图 ------ 邱大发
'''
import os
import openpyxl
from openpyxl.chart import BarChart, Reference, Series
from loggerhelper import *


PATH = os.path.dirname(__file__) #文件运行目录

def get_feedbacker_chart():
    print(PATH)
    wb=openpyxl.load_workbook( PATH + '/total.xlsx')
    ws = wb['total']
    cols=ws['D'] #获取D列反馈人信息
    names = [c.value for c in cols]
    logger.info(names)
    dic = {}
    for n in names:
        name = n.split( )
        if dic.get(name[0],'None')!='None':
            dic[name[0]] += 1
        else:
            dic.setdefault(name[0],1)
        #print(name[0],name[1])
    ss = wb.create_sheet()
    ss.title = 'temp_feedbacker'
    #print(dic.items())
    list_dic = sorted(dic.items(), key=lambda x:x[1],reverse=True )
    #print(listt)
    ss.append(['姓名','次数'])
    for k in list_dic:
        ss.append([k[0],k[1]])
        logger.info([k[0],k[1]])
    add_chart(wb,ss,ws)

    

def add_chart(wb,ss,ws):
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "反馈人"
    chart1.y_axis.title = '次数'
    chart1.x_axis.title = '姓名'
    
    data = Reference(ss, min_col=2, min_row=1,max_row=ss.max_row,max_col=2 )
    cats = Reference(ss, min_col=1, min_row=2,max_col=1, max_row=ss.max_row)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    str = 'B{0}'.format(ws.max_row+10)
    ws.add_chart(chart1,str)
    wb.save(PATH + '/total.xlsx')

