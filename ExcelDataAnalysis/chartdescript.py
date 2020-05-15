'''
对问题描述，做数据分析，输出词云图 ------ 丁小永

'''
import openpyxl
from wordcloud import WordCloud
import jieba
import numpy
import PIL.Image as Image
from openpyxl.drawing.image import Image as IM

from exceloperator import *
from loggerhelper import *

#1.将字符串切分为单个字符8q
def chinese_jieba(text):
    wordlist_jieba=jieba.cut(text)
    space_wordlist=''.join(wordlist_jieba)
    return  space_wordlist

#2.获取tital.xlsx中问题描述的值，返回str
def get_total_data():
    src = os.path.abspath('total.xlsx')
    excelop = ExcelOperator(src)
    problem_desdribe = excelop.get_col_value(7)
    return ''.join(problem_desdribe)

#3.画词云图
def print_wordcloud(text):
    # 1.将字符串切分
    text = chinese_jieba(text)
    # 2.图片遮罩层
    mask_pic = numpy.array(Image.open("pika.jpg"))
    wordcloud = WordCloud(font_path="C:/Windows/Fonts/simfang.ttf",
                          mask=mask_pic).generate(text)
    image = wordcloud.to_image()
    image.save('wordcloud.jpg');
    return image

#4.把词云图输出到excel中
def output_wordcloud():
    wb =  load_workbook("total.xlsx")

    if "wordcloud" in wb.sheetnames:
        sheet = wb.get_sheet_by_name("wordcloud")
    else:
        sheet = wb.create_sheet("wordcloud")

    # 设置文字图片单元格的行高列宽
    column_width = 12.25
    row_height = 80.10

    sheet.column_dimensions['D'].width = column_width
    sheet.row_dimensions[3].height = row_height

    img = IM('wordcloud.jpg')
    newsize = (360, 360)
    img.width, img.height = newsize

    sheet.add_image(img, 'D3')
    wb.save("total.xlsx")  # 保存


#if __name__ == "__main__":
def get_wordcloud():
    text = get_total_data()
    image = print_wordcloud(text)
    output_wordcloud()