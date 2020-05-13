'''
对问题的处理人，做数据分析，输出柱形图 ------ 高国栋
'''
import openpyxl

from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Series, Reference

# 实例化
# wb = workbook()
# 激活worksheet
# ws = wb.active

workbook = load_workbook('total.xlsx')
worksheet = workbook["total"]

# 读取数据部分 存入dict
dict = {}
for cell in list(worksheet.columns)[0]:
    if cell.value in dict:
        dict[cell.value] += 1
    else:
        addtemp = {cell.value: 1}
        dict.update(addtemp)

print("==========测试打印===========")
for eachprint in dict:
    print(eachprint, ":", dict[eachprint])
print("==========测试打印===========")

# 新建工作表 无则创建 有则删除重新创建
all_sheets = workbook.sheetnames
findsheet = False
for i in range(len(all_sheets)):
    if all_sheets[i] == "charthandler":
        findsheet = True
if findsheet:
    worksheetnew = workbook["charthandler"]
    workbook.remove(worksheetnew)
worksheetnew = workbook.create_sheet("charthandler")

# 将数据写入新工作表
worksheetnew.append(["姓名", "处理反馈数"])
for i in dict:
    worksheetnew.append([i, dict[i]])

# 创建表格
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "处理人数据统计"
chart1.y_axis.title = '处理反馈数'
chart1.x_axis.title = '问题处理人'
# row是行 col是列
data = Reference(worksheetnew, min_col=2, min_row=1, max_row=worksheetnew.max_row, max_col=worksheetnew.max_column)
cats = Reference(worksheetnew, min_col=1, min_row=2, max_row=worksheetnew.max_row)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
worksheetnew.add_chart(chart1, "A10")

# 最后保存工作簿
workbook.save("total.xlsx")